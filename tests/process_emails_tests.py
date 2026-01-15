#!/usr/bin/env python
"""Script pour traiter le fichier Excel "Emails type.xlsx", exécuter des tests d'analyse
et ajouter les résultats dans une nouvelle colonne avec la date.
"""

import os
import sys
from datetime import datetime
from typing import Optional

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

# Ajouter le répertoire racine au path pour les imports
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import contextlib

from src.common.config import SupportedLocale  # noqa: E402
from src.common.constants import API_ROUTE_V2  # noqa: E402


def check_backend_health(api_base_url: str) -> bool:
    """Vérifie que le backend est accessible."""
    health_url = f"{api_base_url}/api/health"
    try:
        response = requests.get(health_url, timeout=5.0)
        return response.status_code == 200
    except requests.RequestException:
        return False


def analyze_text(
    api_base_url: str,
    app_id: str,
    locale: SupportedLocale,
    text: str,
    field_values: dict,
    read_from_cache: bool,
    llm_config_id: str,
) -> dict:
    """Appelle l'API d'analyse.

    Returns
    -------
        dict: Résultat de l'analyse ou dict avec 'error' en cas d'erreur

    """
    url = f"{api_base_url}{API_ROUTE_V2}/apps/{app_id}/{locale}/analyze"

    payload = {
        "field_values": field_values,
        "text": text,
        "read_from_cache": read_from_cache,
        "llm_config_id": llm_config_id,
    }

    # Timeout en secondes
    timeout_seconds = 120.0

    try:
        response = requests.post(url, json=payload, timeout=timeout_seconds)

        if response.status_code == 200:
            return response.json()
        else:
            # Pour les erreurs 500, essayer d'extraire plus d'informations
            error_message = response.text[:500]  # Limiter la taille mais plus d'info
            try:
                error_json = response.json()
                if "detail" in error_json:
                    error_message = str(error_json["detail"])
            except ValueError:
                pass

            return {"error": f"HTTP {response.status_code}", "message": error_message}
    except requests.exceptions.ConnectionError:
        return {"error": "connection_error", "message": "Backend inaccessible"}
    except requests.exceptions.Timeout:
        return {
            "error": "timeout",
            "message": f"Timeout lors de l'appel (timeout: {timeout_seconds}s)",
        }
    except Exception as e:
        return {"error": type(e).__name__, "message": str(e)[:200]}


def is_date_column(column_name: str) -> bool:
    """Vérifie si une colonne est une date (format dd/MM/yyyy ou similaire)."""
    if pd.isna(column_name):
        return False

    col_str = str(column_name).strip()
    # Vérifier si c'est un format de date commun
    date_patterns = [
        r"\d{2}/\d{2}/\d{4}",  # dd/MM/yyyy
        r"\d{4}-\d{2}-\d{2}",  # yyyy-MM-dd
        r"\d{2}-\d{2}-\d{4}",  # dd-MM-yyyy
    ]

    import re

    for pattern in date_patterns:
        if re.match(pattern, col_str):
            return True

    # Vérifier si le nom de colonne contient "date" ou "Date"
    if "date" in col_str.lower():
        return True

    return False


def create_field_values(
    row: pd.Series,
    exclude_date_columns: bool = True,
    exclude_columns: Optional[list] = None,
) -> dict:
    """Crée un dictionnaire field_values à partir d'une ligne du DataFrame
    en excluant les colonnes de dates si demandé et les colonnes spécifiées.

    Args:
    ----
        row: Ligne du DataFrame
        exclude_date_columns: Si True, exclut les colonnes de dates
        exclude_columns: Liste des noms de colonnes à exclure

    """
    if exclude_columns is None:
        exclude_columns = []

    field_values = {}

    # Date d'aujourd'hui au format dd/MM/yyyy
    date_demande = datetime.now().strftime("%d/%m/%Y")
    field_values["date_demande"] = date_demande

    # Parcourir toutes les colonnes
    for col in row.index:
        # Exclure les colonnes de dates si demandé
        if exclude_date_columns and is_date_column(col):
            continue

        # Exclure les colonnes spécifiées
        if str(col) in exclude_columns:
            continue

        value = row[col]

        # Ignorer les valeurs NaN
        if pd.isna(value):
            continue

        # Convertir en string et nettoyer
        value_str = str(value).strip()
        if value_str and value_str.lower() != "nan":
            # Utiliser le nom de colonne en minuscules avec underscores
            key = str(col).lower().replace(" ", "_").replace("/", "_")
            field_values[key] = value_str

    return field_values


def process_excel_file(
    input_file: str = "tests/Emails type.xlsx",
    output_file: Optional[str] = None,
    api_base_url: str = "http://localhost:8002",
    app_id: str = "delphes78",
    locale: SupportedLocale = "fr",
    llm_config_id: str = "scaleway1",
    max_rows: Optional[int] = None,
    rows: Optional[str] = None,
) -> None:
    """Traite le fichier Excel, exécute les tests et ajoute les résultats."""
    # Normaliser les chemins relatifs par rapport à la racine du projet
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)

    if not os.path.isabs(input_file):
        input_file = os.path.join(project_root, input_file)

    if output_file is None:
        output_file = input_file
    elif not os.path.isabs(output_file):
        output_file = os.path.join(project_root, output_file)

    # Vérifier que le backend est accessible
    if not check_backend_health(api_base_url):
        return

    # Lire le fichier Excel avec openpyxl pour préserver les styles
    try:
        # Charger avec openpyxl pour préserver les styles
        wb = load_workbook(input_file)
        ws = wb.active

        # Lire aussi avec pandas pour faciliter le traitement des données
        df = pd.read_excel(input_file)
    except Exception:
        return

    # Identifier les colonnes de dates à exclure
    date_columns = [col for col in df.columns if is_date_column(col)]
    if date_columns:
        pass

    # Créer le nom de la nouvelle colonne avec la date actuelle (sera sur 2 lignes)
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    result_column = timestamp

    # Récupérer le modèle LLM pour l'ajouter dans le titre de colonne
    # On va le récupérer lors du premier appel réussi
    llm_model: Optional[str] = None

    # Trouver la colonne de résultats dans le fichier Excel (openpyxl)
    result_col_idx = None
    if result_column in df.columns:
        # La colonne existe déjà dans pandas, trouver son index dans Excel
        col_names = list(df.columns)
        result_col_idx = col_names.index(result_column) + 1  # +1 car Excel commence à 1
    else:
        # Créer une nouvelle colonne
        df[result_column] = None
        # La nouvelle colonne sera directement après la dernière colonne existante dans Excel
        # Utiliser ws.max_column pour obtenir la vraie dernière colonne du fichier Excel
        result_col_idx = ws.max_column + 1

    # Colonne pour le texte à analyser
    text_column = "Texte Mail"
    if text_column not in df.columns:
        return

    # Sauvegarder le nombre total de lignes
    total_rows = len(df)

    # Limiter le nombre de lignes à traiter si demandé (mais garder toutes les lignes dans le DataFrame)
    rows_to_process = df
    row_numbers_list = None
    if rows:
        # Traiter seulement les lignes spécifiées (1-indexed dans l'argument, converties en 0-indexed pour le DataFrame)
        row_numbers_list = [int(x.strip()) for x in rows.split(",")]
        # Convertir de 1-indexed (lignes Excel) à 0-indexed (indices DataFrame)
        # Ligne Excel 3 = DataFrame index 1 (car ligne 1 = header, ligne 2 = index 0)
        df_indices = [
            row_num - 2
            for row_num in row_numbers_list
            if 2 <= row_num <= total_rows + 1
        ]
        if df_indices:
            # Sélectionner les lignes par position
            rows_to_process = df.iloc[df_indices].copy()
        else:
            rows_to_process = df
            row_numbers_list = None
    elif max_rows is not None:
        rows_to_process = df.head(max_rows)

    # Traiter chaque ligne
    len(rows_to_process)

    # Gestion des interruptions pour fermer proprement le fichier Excel
    try:
        # Utiliser enumerate pour avoir la position dans rows_to_process (0-indexed)
        # Mais on doit trouver la vraie ligne Excel correspondante
        for pos, (idx, row) in enumerate(rows_to_process.iterrows()):
            # Timestamp
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Composer le message à partir des colonnes F (Objet du mail) et G (Texte Mail)
            objet_column = "Objet du mail"
            texte_column = "Texte Mail"

            # Récupérer l'objet et le texte
            objet = (
                str(row[objet_column])
                if objet_column in row.index and not pd.isna(row[objet_column])
                else ""
            )
            texte = (
                str(row[texte_column])
                if texte_column in row.index and not pd.isna(row[texte_column])
                else ""
            )

            # Supprimer "[INTERNET] " au début de l'objet
            if objet.startswith("[INTERNET] "):
                objet = objet[11:]  # Supprimer "[INTERNET] "

            # Composer le message : objet + 2 sauts de ligne + texte
            text_str = f"{objet}\n\n{texte}".strip()

            # Calculer le numéro de ligne Excel pour l'affichage
            if rows and row_numbers_list and pos < len(row_numbers_list):
                # Utiliser directement la liste row_numbers_list avec la position
                row_numbers_list[pos]
            else:
                pos + 2

            # Ignorer les lignes sans texte
            if not text_str.strip():
                continue

            # Trace : Timestamp et numéro de ligne

            # Trace : Message composé (limité à 200 caractères pour l'affichage)
            (text_str[:200] + "..." if len(text_str) > 200 else text_str)

            # Créer le payload (exclure les colonnes de dates et "Type de problème")
            exclude_cols = ["Type de problème"]
            field_values = create_field_values(
                row,
                exclude_date_columns=True,
                exclude_columns=exclude_cols,
            )

            # Appeler l'API avec gestion d'erreur
            is_error = False
            result = None
            import time

            start_time = time.time()
            try:
                # Afficher un indicateur de progression pendant l'attente
                result = analyze_text(
                    api_base_url=api_base_url,
                    app_id=app_id,
                    locale=locale,
                    text=text_str,
                    field_values=field_values,
                    read_from_cache=False,
                    llm_config_id=llm_config_id,
                )
                time.time() - start_time
            except Exception as e:
                # En cas d'exception, créer un résultat d'erreur
                result = {"error": "exception", "message": str(e)}
                is_error = True
                result_str = f"ERREUR: Exception - {e!s}"

            # Stocker le résultat
            result_str = ""
            if not is_error:
                if "error" in result:
                    result_str = f"ERREUR: {result.get('error', 'unknown')} - {result.get('message', '')}"
                    is_error = True
                else:
                    # Formater le résultat de manière détaillée
                    result_lines = []

                    if "analysis_result" in result:
                        analysis = result["analysis_result"]

                    # Récupérer le temps de réponse et le modèle LLM
                    response_time = "N/A"
                    if "statistics" in analysis:
                        stats = analysis["statistics"]
                        response_time_str = stats.get("Response time", "N/A")
                        # Enlever le "s" à la fin si présent
                        if isinstance(
                            response_time_str,
                            str,
                        ) and response_time_str.endswith("s"):
                            response_time = response_time_str[:-1]
                        else:
                            response_time = response_time_str

                        # Récupérer le modèle LLM pour le titre de colonne (première fois)
                        if llm_model is None:
                            llm_model = stats.get("LLM Model", "N/A")

                    # Intentions avec temps de réponse dans le titre
                    if analysis.get("scorings"):
                        result_lines.append(
                            f"Intentions (temps de réponse : {response_time} secondes) :",
                        )

                        # Filtrer les intentions avec score > 0, puis trier par score décroissant
                        filtered_scorings = [
                            s for s in analysis["scorings"] if s.get("score", 0) > 0
                        ]
                        sorted_scorings = sorted(
                            filtered_scorings,
                            key=lambda x: x.get("score", 0),
                            reverse=True,
                        )

                        if sorted_scorings:
                            for idx, scoring in enumerate(sorted_scorings, 1):
                                intention_label = scoring.get("intention_label", "N/A")
                                score = scoring.get("score", 0)
                                justification = scoring.get("justification", "N/A")

                                result_lines.append(f"{idx}/ {intention_label}")
                                result_lines.append(f"   - Score: {score}")
                                result_lines.append(
                                    f"   - Justification: {justification}",
                                )
                        else:
                            result_lines.append("Aucune intention avec score > 0")
                    else:
                        result_lines.append(
                            f"Intentions (temps de réponse : {response_time} secondes) :",
                        )
                        result_lines.append("Aucune intention trouvée")

                    result_str = "\n".join(result_lines)

                    # Trace : Temps de réponse
                    if "statistics" in analysis:
                        stats = analysis["statistics"]
                        response_time_str = stats.get("Response time", "N/A")
                        if isinstance(
                            response_time_str,
                            str,
                        ) and response_time_str.endswith("s"):
                            pass
                        else:
                            pass
                    else:
                        pass

                    # Trace : Résultat (intention principale)
                    if analysis.get("scorings"):
                        top_intention = max(
                            analysis["scorings"],
                            key=lambda x: x.get("score", 0),
                        )
                        intention_label = top_intention.get("intention_label", "N/A")
                        score = top_intention.get("score", 0)
                    else:
                        pass
            else:
                result_str = "Résultat inattendu"
                is_error = True

            # Stocker le résultat directement dans le fichier Excel avec openpyxl pour préserver les styles
            # Si on a spécifié des lignes avec --rows, utiliser directement row_numbers_list[pos]
            # Sinon, pos est la position dans rows_to_process (0-indexed)
            if rows and row_numbers_list and pos < len(row_numbers_list):
                # Utiliser directement la liste row_numbers_list avec la position
                excel_row = row_numbers_list[pos]
            else:
                # Utiliser directement pos + 2 car rows_to_process contient les lignes dans l'ordre
                excel_row = pos + 2

            try:
                # Copier le style d'un en-tête existant (première colonne par exemple)
                header_font = None
                header_fill = None
                header_border = None
                if ws.max_column > 0:
                    # Prendre le style de la première colonne d'en-tête comme référence
                    ref_header = ws.cell(row=1, column=1)
                    if ref_header.font:
                        header_font = Font(
                            name=ref_header.font.name or "Calibri",
                            size=ref_header.font.size or 11,
                            bold=(
                                ref_header.font.bold
                                if ref_header.font.bold is not None
                                else True
                            ),
                            color=ref_header.font.color,
                        )
                    if ref_header.fill:
                        header_fill = PatternFill(
                            fill_type=ref_header.fill.fill_type or "solid",
                            start_color=ref_header.fill.start_color or "FFFFFF",
                            end_color=ref_header.fill.end_color or "FFFFFF",
                        )
                    # Copier la bordure en créant un nouvel objet Border
                    if ref_header.border:
                        try:
                            header_border = Border(
                                left=ref_header.border.left,
                                right=ref_header.border.right,
                                top=ref_header.border.top,
                                bottom=ref_header.border.bottom,
                            )
                        except Exception:
                            # Si erreur, on ne copie pas la bordure
                            header_border = None

                # Si c'est une nouvelle colonne, créer l'en-tête
                if result_col_idx > ws.max_column:
                    header_cell = ws.cell(row=1, column=result_col_idx)
                    header_cell.value = (
                        timestamp  # Temporaire, sera mis à jour avec le modèle LLM
                    )

                    # Appliquer le style de l'en-tête
                    if header_font:
                        header_cell.font = header_font
                    if header_fill:
                        header_cell.fill = header_fill
                    if header_border:
                        header_cell.border = header_border
                    header_cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="center",
                    )

                    ws.row_dimensions[1].height = 50

                    # Définir la largeur de la colonne à 350 pixels (environ 50 unités Excel)
                    # 1 unité Excel ≈ 7 pixels, donc 350 pixels ≈ 50 unités
                    col_letter = get_column_letter(result_col_idx)
                    ws.column_dimensions[col_letter].width = 50

                # Écrire le résultat dans la cellule Excel (toujours écrire quelque chose)
                result_cell = ws.cell(row=excel_row, column=result_col_idx)
                if result_str and result_str.strip():
                    result_cell.value = result_str
                else:
                    result_cell.value = ""  # Écrire au moins une chaîne vide
                result_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Appliquer le style
                if is_error:
                    # Pour les erreurs : texte en rouge, fond blanc transparent
                    result_cell.font = Font(
                        name="Calibri",
                        size=11,
                        color=Color(rgb="00FF0000"),  # Rouge
                    )
                    result_cell.fill = PatternFill(
                        fill_type="solid",
                        start_color="FFFFFF",
                        end_color="FFFFFF",
                    )  # Fond blanc
                elif excel_row <= ws.max_row and ws.max_column > 0:
                    # Prendre le style de la première colonne de la même ligne
                    ref_cell = ws.cell(row=excel_row, column=1)

                    # Copier la police (font) - important pour la couleur du texte
                    if ref_cell.font:
                        try:
                            result_cell.font = Font(
                                name=ref_cell.font.name or "Calibri",
                                size=ref_cell.font.size or 11,
                                bold=ref_cell.font.bold,
                                italic=ref_cell.font.italic,
                                underline=ref_cell.font.underline,
                                strike=ref_cell.font.strike,
                                color=ref_cell.font.color,  # Copier la couleur du texte
                            )
                        except Exception:
                            # Style par défaut si erreur
                            result_cell.font = Font(name="Calibri", size=11)
                    else:
                        result_cell.font = Font(name="Calibri", size=11)

                    # Copier la bordure
                    if ref_cell.border:
                        with contextlib.suppress(Exception):
                            result_cell.border = Border(
                                left=ref_cell.border.left,
                                right=ref_cell.border.right,
                                top=ref_cell.border.top,
                                bottom=ref_cell.border.bottom,
                            )
                else:
                    # Style par défaut si pas de référence
                    result_cell.font = Font(name="Calibri", size=11)

                # Toujours appliquer un fond blanc transparent (sauf pour les erreurs qui ont déjà un fond blanc)
                if not is_error:
                    result_cell.fill = PatternFill(
                        fill_type="solid",
                        start_color="FFFFFF",
                        end_color="FFFFFF",
                    )  # Fond blanc transparent

                # Mettre à jour le titre de colonne si on a le modèle LLM
                if llm_model:
                    header_cell = ws.cell(row=1, column=result_col_idx)
                    header_cell.value = f"{timestamp}\n{llm_model}"

                    # Réappliquer le style
                    if header_font:
                        header_cell.font = header_font
                    if header_fill:
                        header_cell.fill = header_fill
                    if header_border:
                        header_cell.border = header_border
                    header_cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="center",
                    )

                    ws.row_dimensions[1].height = 50

                    # S'assurer que la largeur de colonne est définie
                    col_letter = get_column_letter(result_col_idx)
                    ws.column_dimensions[col_letter].width = 50

                # Sauvegarder après chaque ligne pour ne pas perdre les résultats
                wb.save(output_file)

                # Mettre à jour aussi le DataFrame pour la cohérence
                df.at[idx, result_column] = result_str

            except Exception:
                pass

    except KeyboardInterrupt:
        with contextlib.suppress(Exception):
            wb.save(output_file)
        return

    # Sauvegarde finale - le fichier a déjà été sauvegardé ligne par ligne
    # On s'assure juste que le titre de colonne est bien formaté et que la largeur est définie
    try:
        # Le fichier a déjà été sauvegardé ligne par ligne avec openpyxl
        # On s'assure juste que le titre de colonne est correct et que la largeur est définie
        if result_col_idx:
            # Copier le style d'un en-tête existant
            header_font = None
            header_fill = None
            header_border = None
            if ws.max_column > 0:
                ref_header = ws.cell(row=1, column=1)
                if ref_header.font:
                    header_font = Font(
                        name=ref_header.font.name or "Calibri",
                        size=ref_header.font.size or 11,
                        bold=(
                            ref_header.font.bold
                            if ref_header.font.bold is not None
                            else True
                        ),
                        color=ref_header.font.color,
                    )
                if ref_header.fill:
                    header_fill = PatternFill(
                        fill_type=ref_header.fill.fill_type or "solid",
                        start_color=ref_header.fill.start_color or "FFFFFF",
                        end_color=ref_header.fill.end_color or "FFFFFF",
                    )
                # Copier la bordure en créant un nouvel objet Border
                if ref_header.border:
                    try:
                        header_border = Border(
                            left=ref_header.border.left,
                            right=ref_header.border.right,
                            top=ref_header.border.top,
                            bottom=ref_header.border.bottom,
                        )
                    except Exception:
                        # Si erreur, on ne copie pas la bordure
                        header_border = None

            header_cell = ws.cell(row=1, column=result_col_idx)
            if llm_model:
                header_cell.value = f"{timestamp}\n{llm_model}"
            else:
                header_cell.value = timestamp

            # Appliquer le style
            if header_font:
                header_cell.font = header_font
            if header_fill:
                header_cell.fill = header_fill
            if header_border:
                header_cell.border = header_border
            header_cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center",
            )

            ws.row_dimensions[1].height = 50

            # Définir la largeur de la colonne à 350 pixels (50 unités Excel)
            col_letter = get_column_letter(result_col_idx)
            ws.column_dimensions[col_letter].width = 50

            wb.save(output_file)

    except Exception:
        return


def main() -> None:
    """Fonction principale."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Traite un fichier Excel et exécute des tests d'analyse",
    )
    parser.add_argument(
        "--input",
        default="tests/Emails type.xlsx",
        help="Fichier Excel d'entrée (relatif à la racine du projet)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Fichier Excel de sortie (par défaut: même que l'entrée)",
    )
    parser.add_argument(
        "--api-url",
        default="http://localhost:8002",
        help="URL de l'API backend",
    )
    parser.add_argument("--app-id", default="delphes78", help="ID de l'application")
    parser.add_argument("--locale", default="fr", help="Locale")
    parser.add_argument("--llm-config", default="scaleway1", help="Configuration LLM")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Nombre maximum de lignes à traiter (pour les tests)",
    )
    parser.add_argument(
        "--rows",
        type=str,
        default=None,
        help="Liste de numéros de lignes Excel à traiter (1-indexed, séparés par des virgules, ex: '3,12,27')",
    )

    args = parser.parse_args()

    process_excel_file(
        input_file=args.input,
        output_file=args.output,
        api_base_url=args.api_url,
        app_id=args.app_id,
        locale=args.locale,
        llm_config_id=args.llm_config,
        max_rows=args.max_rows,
        rows=args.rows,
    )


if __name__ == "__main__":
    main()
