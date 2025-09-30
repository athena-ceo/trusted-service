import json
from typing import Any, Literal, Type

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from pydantic import BaseModel, field_validator

# IF YOU CHANGE THE FOLLOWING COMMENT, UPDATE README.md ACCORDINGLY
# Add here support for new languages
SupportedLocale = Literal["en", "fr", "fi", "sv"]


class Config(BaseModel):
    pass


def load_dicts_from_worksheet(worksheet, locale: SupportedLocale) -> list[dict[str, Any]]:
    dicts: list[dict[str, Any]] = []

    title_row_hit = False
    column_labels_and_numbers: dict[str, int] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        # Empty rows
        if row[0].value is None or str(row[0].value).startswith("#"):
            continue

        # Header row
        if not title_row_hit:
            title_row_hit = True
            for cell in row:
                title = str(cell.value)
                if title.endswith(f"_{locale}"):
                    title = title[:len(title) - len(locale) - 1]
                column_labels_and_numbers[title] = cell.column
            continue

        # Value rows
        data: dict[str, Any] = {}
        for label, number in column_labels_and_numbers.items():
            data[label] = row[number - 1].value
        dicts.append(data)
    return dicts


def load_pydantic_objects_from_worksheet(worksheet, model_type: type[BaseModel], locale: SupportedLocale) -> list[BaseModel]:
    list1: list[dict[str, Any]] = load_dicts_from_worksheet(worksheet, locale)
    return [model_type.model_validate(data) for data in list1]


def load_pydantic_objects_from_worksheet2(worksheet, model_type: type[BaseModel], locale: SupportedLocale) -> list[BaseModel]:
    objects = []

    title_row_hit = False
    column_labels_and_numbers: dict[str, int] = {}
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        # Empty rows
        if row[0].value is None or str(row[0].value).startswith("#"):
            continue

        # Header row
        if not title_row_hit:
            title_row_hit = True
            for cell in row:
                title = str(cell.value)
                if title.endswith(f"_{locale}"):
                    title = title[:len(title) - len(locale) - 1]
                column_labels_and_numbers[title] = cell.column
            continue

        # Value rows
        data: dict[str, Any] = {}
        for label, number in column_labels_and_numbers.items():
            data[label] = row[number - 1].value
        # print("data", data)
        obj = model_type.model_validate(data)
        objects.append(obj)
    return objects


def load_config_from_workbook(filename: str,
                              main_tab: str | None,
                              collections: list[tuple[str, type[BaseModel]]],
                              config_type: Type[Config],
                              locale: SupportedLocale | None) -> Config:
    config_workbook: Workbook = load_workbook(filename)
    config_values: dict[str, Any] = {}

    # main_tab
    if main_tab:
        worksheet = config_workbook[main_tab]

        title_row_hit = False
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            if row[0].value is None:
                continue
            if not title_row_hit:
                title_row_hit = True
                continue
            key = row[0].value
            if key.endswith(f"_{locale}"):
                key = key[:len(key) - len(locale) - 1]
            config_values[key] = row[1].value

    for collection_name, model_type in collections:
        config_values[collection_name] = load_pydantic_objects_from_worksheet(
            worksheet=config_workbook[collection_name],
            model_type=model_type,
            locale=locale)

    return config_type.model_validate(config_values)
